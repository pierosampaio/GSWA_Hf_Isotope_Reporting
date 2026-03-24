import pandas as pd
import numpy as np
from scipy.stats import norm
import re

import os
import sys
from pathlib import Path



script_path = r"C:\workplace\PythonScripts"
if script_path not in sys.path:
    sys.path.append(script_path)

from definitions import *
from functions import *

from scipy.optimize import brentq

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment






def path_checker(Prompt):
  path = input(Prompt)
  if not Path(path).is_dir():
    print("Directory does not exist. Try again.")
    return path_checker(Prompt)
  else:
    return path



def main():
  
  UPb_path = path_checker("U-Pb files directory:")
  LuHf_file = input("Lu-Hf file path:")
  
    
  
  UPb_files = os.listdir(UPb_path)
  
  out_path = input("Output directory:")
  if not Path(out_path).is_dir():
    os.mkdir(out_path)
    
  
  UPb = UPb_file_join(UPb_path)
  try:
    LuHf = LuHf_process(LuHf_file, Sample_list = UPb_files)
  except NameError:
    print("File does not exist.")
    
  df_merged = merge_datasets(UPb,LuHf,"SampleSpot")
  df_merged = df_merged.drop("Sample_y", axis = 1)
  df_merged = df_merged.rename({
    "Sample_x": "Sample"
  }, axis = 1)
  

  print(f"Number of matched spots: {len(df_merged)}")
  
  ############## Calculate weighted mean ages for the igneous crystals
  
  df_merged_filtered = df_merged.loc[
    df_merged.GroupID.isin([
      "I"
    ])
  ]
   
  sample_weighted_mean = (
    df_merged_filtered
    .groupby("Sample")
    .apply(lambda g: weighted_mean(
      g["4corr_Pb207Pb206"],
      g["4corr_Pb207Pb206_1sig"],
      method = "chauvenet"
    ))
  )
  
  sample_weighted_mean = pd.DataFrame.from_dict(
    sample_weighted_mean.to_dict(),
    orient = "index"
  )
  
  sample_weighted_mean[["age","sigma_age"]] = (
    sample_weighted_mean.apply(
      lambda s: pb207_pb206_age_with_uncertainty(
        s["mean"], 2*s["uncertainty"]
      ), axis =1, result_type = "expand"
    )
  )
  
  df_merged.loc[df_merged.GroupID.isin(["I"]), "Age_Hf_calculation"] = \
    df_merged.loc[df_merged.GroupID.isin(["I"]), "Sample"].map(
      sample_weighted_mean["age"]
    )
    
  df_merged.loc[df_merged.GroupID.isin(["I"]), "Age_Hf_calculation_unc"] = \
    df_merged.loc[df_merged.GroupID.isin(["I"]), "Sample"].map(
      sample_weighted_mean["sigma_age"]
    )
    
  df_merged.loc[~df_merged.GroupID.isin(["I"]), "Age_Hf_calculation"] = \
    df_merged.loc[~df_merged.GroupID.isin(["I"]), "4corr_76_date"]
  
  df_merged.loc[~df_merged.GroupID.isin(["I"]), "Age_Hf_calculation_unc"] = \
    df_merged.loc[~df_merged.GroupID.isin(["I"]), "4corr_76_date_1sig"]
  
  ################################################################################
  
  
  overwrite = input("Do you want to overwrite any of the igneous ages for isotope ratio back-calculations (Y/N)?:")
  if overwrite.upper() == "Y":
    df_merged = overwrite_igneous_age(df_merged)
    
  else: 
    pass
  
  
  #### Bit botched, have to check how to proceed regarding these uncertainties
  #df_merged["Hf176Hf177_1sig"] = df_merged["Hf176Hf177_2SE"]
  df_merged["Hf176Hf177_1SE"] = df_merged["Hf176Hf177_2SE"]/2

  
  df_merged["Hf176Hf177_init"] = df_merged.apply(
      lambda s: calc_initial_ratios(
          s["Hf176Hf177"],
          s["Lu176Hf177"],
          s["Age_Hf_calculation"]
      ), axis = 1
  )
  
  df_merged["Hf176Hf177_init_1SE"] = df_merged.apply(
      lambda s: initial_ratio_uncertainty(
          s["Hf176Hf177_1SE"]
      ), axis = 1
  )
  
  df_merged["epsilon_Hf"] = df_merged.apply(
      lambda s: epsilon_Hf(
          s["Hf176Hf177"],
          s["Lu176Hf177"],
          s["Age_Hf_calculation"]
      ), axis = 1
  )
  
  df_merged["epsilon_Hf_1SE"] = df_merged.apply(
      lambda s: epsilon_Hf_uncertainty(
          s["Hf176Hf177_1SE"],
          s["Age_Hf_calculation"]
      ), axis = 1
  )
  
  df_merged["TDM_2Stage"] = df_merged.apply(
      lambda s: two_stage_TDM_YJ(
          s["Hf176Hf177_init"],
          s["Lu176Hf177"],
          s["Age_Hf_calculation"]
      ), axis = 1
  ).round(0)
  
  df_merged["TDM_2Stage"]
  
  df_merged["TDM_2Stage_5pctUnc"] = np.round(df_merged["TDM_2Stage"].values * 0.05,0)
  df_merged["TCR"] = (df_merged["TDM_2Stage"].values - df_merged["Age_Hf_calculation"].values).round(0)
  df_merged["TCR_5pctUnc"] = np.round(df_merged["TCR"].values * 0.05, 0)
  
 

###############################################################################
################################ WAGIMS_EXPORT ################################

  col_order = [
      "Session date","Sample", "UPB_ANALYSIS_ID","GroupID","238U_ppm", "232Th_ppm", "Th232U238", "f204_pct",
      "U238Pb206", "U238Pb206_1sig", "Pb207Pb206", "Pb207Pb206_1sig", 
      "4corr_U238Pb206", "4corr_U238Pb206_1sig", "4corr_Pb207Pb206", "4corr_Pb207Pb206_1sig",
      "4corr_86_date","4corr_86_date_1sig","4corr_76_date","4corr_76_date_1sig",
      "Discordance_pct","Material analysed","LU_HF_ANALYSIS_ID","Age_Hf_calculation",
      "Age_Hf_calculation_unc","Hf176Hf177","Hf176Hf177_2SE","Hf178Hf177","Hf178Hf177_2SE",
      "Lu176Hf177","Lu176Hf177_2SE","Yb176Hf177","Yb176Hf177_2SE","Comment",
      "Hf176Hf177_init","Hf176Hf177_init_1SE","epsilon_Hf","epsilon_Hf_1SE","TDM_2Stage",
      "TCR"
  ]
  
  WAGIMS_names = [                   # following WAGIMS nomenclature
      "Session date", "Sample ID", "Geochron analysis ID", "Geochron group ID", "238U (ppm)", "232Th (ppm)",
      "232Th/238U", "f204 (%)", "238U/206Pb", "238U/206Pb 1 sigma", "207Pb/206Pb", "207Pb/206Pb 1 sigma",
      "238U/206Pb*", "238U/206Pb* 1 sigma", "207Pb*/206Pb*", "207Pb*/206Pb* 1 sigma",
      "204-corr 206Pb/238U date (Ma)", "204-corr 206Pb/238U date 1 sigma (Ma)",
      "204-corr 207Pb/206Pb date (Ma)", "204-corr 207Pb/206Pb date 1 sigma (Ma)",
      "204-corr 6-38 v 7-6 discordance (%)","Material analysed","Lu-Hf analysis ID",
      "Age for initial Hf calculation (Ma)", "Age for initial Hf calculation 1 sigma (Ma)",
      "176Hf/177Hf measured", "176Hf/177Hf measured 2SE", "178Hf/177Hf measured",
      "178Hf/177Hf measured 2SE", "176Lu/177Hf measured", "176Lu/177Hf measured 2SE",
      "176Yb/177Hf measured", "176Yb/177Hf measured 1 sigma", "Comment", "176Hf/177Hf initial",
      "176Hf/177Hf initial 1SE", "Epsilon Hf(i)", "Epsilon Hf(i) 1SE", "TDM2 (Ma)",
      "TCR (Ma)"
  ]
  
  df_merged.insert(len(df_merged.columns), "Material analysed","zircon")
  df_merged.insert(len(df_merged.columns), "Comment","")
  df_merged.insert(0,"Session date","")
  
  
  
  df_merged_WAGIMS = df_merged[col_order]
  
  df_merged_WAGIMS = df_merged_WAGIMS.rename(
      dict(zip(col_order,WAGIMS_names)), axis = 1)
  
  df_merged_WAGIMS["Sample ID"] = df_merged_WAGIMS["Sample ID"].str.split(".", expand = True)[0]
  df_merged_WAGIMS["Geochron analysis ID"] = df_merged_WAGIMS["Geochron analysis ID"].astype("str")
  
  
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"Laboratory","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"Instrument","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"Laser system","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"Pulse width","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"Laser wavelength (nm)","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"LASS (yes/no)","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"Laser spot size (μm)","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"Laser fluence (J/cm2)","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"Repetition rate (Hz)","")
  
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"RM1","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"RM1 mean 176Hf/177Hf","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"RM1 mean 176Hf/177Hf 2 SD","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"RM2","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"RM2 mean 176Hf/177Hf","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"RM2 mean 176Hf/177Hf 2 SD","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"RM3","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"RM3 mean 176Hf/177Hf","")
  df_merged_WAGIMS.insert(len(df_merged_WAGIMS.columns),"RM3 mean 176Hf/177Hf 2 SD","")
  
  Group1 = ["Session"]
  Group2 = ["Sample ID"]
  Group3 = ["Geochron analysis ID", "Geochron group ID", "238U (ppm)", "232Th (ppm)",
      "232Th/238U", "f204 (%)", "238U/206Pb", "238U/206Pb 1 sigma", "207Pb/206Pb", "207Pb/206Pb 1 sigma",
      "238U/206Pb*", "238U/206Pb* 1 sigma", "207Pb*/206Pb*", "207Pb*/206Pb* 1 sigma",
      "204-corr 206Pb/238U date (Ma)", "204-corr 206Pb/238U date 1 sigma (Ma)",
      "204-corr 207Pb/206Pb date (Ma)", "204-corr 207Pb/206Pb date 1 sigma (Ma)",
      "204-corr 6-38 v 7-6 discordance (%)"]
  Group4 = [
      "Material analysed","Lu-Hf analysis ID",
      "Age for initial Hf calculation (Ma)", "Age for initial Hf calculation 1 sigma (Ma)",
      "176Hf/177Hf measured", "176Hf/177Hf measured 1 sigma", "178Hf/177Hf measured",
      "178Hf/177Hf measured sigma", "176Lu/177Hf measured", "176Lu/177Hf measured 1 sigma",
      "176Yb/177Hf measured", "176Yb/177Hf measured 1 sigma", "Comment"
  ]
  Group5 = [
      "176Hf/177Hf initial","176Hf/177Hf initial 1 sigma",
      "Epsilon Hf(i)", "Epsilon Hf(i) 1 sigma", "TDM2 (Ma)",
      "TCR (Ma)"
  ]
  Group6 = [
      "Laboratory","Instrument","Laser system","Pulse width","Laser wavelength (nm)",
      "LASS (yes/no)","Laser spot size (μm)","Laser fluence (J/cm2)",
      "Repetition rate (Hz)"
  ]
  Group7 = [
      "RM1","RM1 mean 176Hf/177Hf","RM1 mean 176Hf/177Hf 2 SD",
      "RM2","RM2 mean 176Hf/177Hf","RM2 mean 176Hf/177Hf 2 SD",
      "RM3","RM3 mean 176Hf/177Hf","RM3 mean 176Hf/177Hf 2 SD"
  ]
  
  
  Higher_order = [
      "SESSION", "SAMPLE INFORMATION", "GEOCHRONOLOGY DATA", "Lu-Hf ANALYSES", "CALCULATED VALUES",
      "LABORATORY INFORMATION", "REFERENCE MATERIALS (RM)"
  ]
  Lower_order = [
      Group1,Group2,Group3,Group4,Group5,Group6,Group7
  ]
  
  
  MultiIndex = []
  for h,l in zip(Higher_order,Lower_order):
      for c in l:
          MultiIndex.append((h,c))
  
  MultiIndex
  
  
  # Multi-level column structure
  columns = pd.MultiIndex.from_tuples(
      MultiIndex
  )
  df = pd.DataFrame(df_merged_WAGIMS.values,columns=columns)


  styled = (
    df.style
    # ---- LEVEL 0 (top header groups) ----
    .applymap_index(
        lambda v: (
            'background-color: #FBE5D6; font-weight: bold'
            if v == 'SAMPLE INFORMATION' else
            'background-color: #FFF2CC; font-weight: bold'
            if v == 'GEOCHRONOLOGY DATA' else
            'background-color: #E2F0D9; font-weight: bold'
            if v == 'SESSION' else
            'background-color: #FFFFCC; font-weight: bold'
            if v == 'Lu-Hf ANALYSES' else
            'background-color: #DEEBF7; font-weight: bold'
            if v == 'CALCULATED VALUES' else
            'background-color: #DEEBF7; font-weight: bold'
            if v == 'LABORATORY INFORMATION' else
            'background-color: #E2F0D9; font-weight: bold'
            if v == 'REFERENCE MATERIALS (RM)' else            
            ''
        ),
        axis=1,
        level=0
    )

    # ---- LEVEL 1 (sub-columns) ----
    .applymap_index(
        lambda v: (
            'background-color: #A9D18E; font-weight: bold'
            if v in Group1 else
            'background-color: #F8CBAD; font-weight: bold'
            if v in Group2 else
            'background-color: #FFE699; font-weight: bold'
            if v in Group3 else
            'background-color: #FFFF99; font-weight: bold'
            if v in Group4 else
            'background-color: #BDD7EE; font-weight: bold'
            if v in Group5 else
            'background-color: #B4C7E7; font-weight: bold'
            if v in Group6 else
            'background-color: #C5E0B4; font-weight: bold'
            if v in Group7 else
            ''
        ),
        axis=1,
        level=1
    )
)

  format_1dp = '0.0'
  format_2dp = '0.00'
  format_3dp = '0.000'
  format_4dp = '0.0000' 
  format_5dp = '0.00000'
  format_6dp = '0.000000'




  spot_level_name = input("Spot-level file name (include extension):") 
  spot_level = os.path.join(out_path,spot_level_name)

  with pd.ExcelWriter(spot_level, engine="openpyxl") as writer:
    styled.to_excel(writer, sheet_name="Lu-Hf")

    workbook  = writer.book
    worksheet = writer.sheets["Lu-Hf"]

    ######### row heights

    for cell in worksheet[1]:
        cell.alignment = Alignment(wrap_text=True)
    for cell in worksheet[2]:
        cell.alignment = Alignment(wrap_text=True)


    ######### column widths
    for i, col in enumerate(df_merged_WAGIMS.columns, 2):

        if get_column_letter(i) == "B":
            worksheet.column_dimensions["B"].width = 10.3
        if get_column_letter(i) == "C":
            worksheet.column_dimensions["C"].width = 15
        if get_column_letter(i) == "D":
            worksheet.column_dimensions["D"].width = 10.71
        if get_column_letter(i) == "E":
            worksheet.column_dimensions["E"].width = 9
        if get_column_letter(i) == "F":
            worksheet.column_dimensions["F"].width = 6.6
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "G":
            worksheet.column_dimensions["G"].width = 6.6
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "H":
            worksheet.column_dimensions["H"].width = 6.6
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_2dp
        if get_column_letter(i) == "I":
            worksheet.column_dimensions["I"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_3dp
        if get_column_letter(i) == "I":
            worksheet.column_dimensions["I"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_3dp
        if get_column_letter(i) == "J":
            worksheet.column_dimensions["J"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_3dp
        if get_column_letter(i) == "K":
            worksheet.column_dimensions["K"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_3dp
        if get_column_letter(i) == "L":
            worksheet.column_dimensions["L"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_5dp
        if get_column_letter(i) == "M":
            worksheet.column_dimensions["M"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_5dp
        if get_column_letter(i) == "N":
            worksheet.column_dimensions["N"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_3dp
        if get_column_letter(i) == "O":
            worksheet.column_dimensions["O"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_3dp
        if get_column_letter(i) == "P":
            worksheet.column_dimensions["P"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_5dp
        if get_column_letter(i) == "Q":
            worksheet.column_dimensions["Q"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_5dp
        if get_column_letter(i) == "R":
            worksheet.column_dimensions["R"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "S":
            worksheet.column_dimensions["S"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "T":
            worksheet.column_dimensions["T"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "U":
            worksheet.column_dimensions["U"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "V":
            worksheet.column_dimensions["V"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_2dp
        if get_column_letter(i) == "X":
            worksheet.column_dimensions["X"].width = 13
        if get_column_letter(i) == "Y":
            worksheet.column_dimensions["Y"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "Z":
            worksheet.column_dimensions["Z"].width = 8.5
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "AA":
            worksheet.column_dimensions["AA"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "AB":
            worksheet.column_dimensions["AB"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "AC":
            worksheet.column_dimensions["AC"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "AD":
            worksheet.column_dimensions["AD"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "AE":
            worksheet.column_dimensions["AE"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "AF":
            worksheet.column_dimensions["AF"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "AG":
            worksheet.column_dimensions["AG"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_4dp
        if get_column_letter(i) == "AH":
            worksheet.column_dimensions["AH"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_4dp
        if get_column_letter(i) == "AJ":
            worksheet.column_dimensions["AJ"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "AK":
            worksheet.column_dimensions["AK"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "AL":
            worksheet.column_dimensions["AL"].width = 9.4
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_2dp
        if get_column_letter(i) == "AM":
            worksheet.column_dimensions["AM"].width = 9.4
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_2dp
        if get_column_letter(i) == "AN":
            worksheet.column_dimensions["AN"].width = 9.4
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "AO":
            worksheet.column_dimensions["AO"].width = 9.4
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = "0"
        if get_column_letter(i) == "AP":
            worksheet.column_dimensions["AP"].width = 21
        if get_column_letter(i) == "AQ":
            worksheet.column_dimensions["AQ"].width = 34
        if get_column_letter(i) == "AR":
            worksheet.column_dimensions["AR"].width = 28
        if get_column_letter(i) == "AS":
            worksheet.column_dimensions["AS"].width = 12
        if get_column_letter(i) == "AT":
            worksheet.column_dimensions["AT"].width = 15
        if get_column_letter(i) == "AU":
            worksheet.column_dimensions["AU"].width = 9
        if get_column_letter(i) == "AV":
            worksheet.column_dimensions["AV"].width = 8.15
        if get_column_letter(i) == "AW":
            worksheet.column_dimensions["AW"].width = 8.15
        if get_column_letter(i) == "AX":
            worksheet.column_dimensions["AX"].width = 8.15
        if get_column_letter(i) == "AY":
            worksheet.column_dimensions["AY"].width = 9.5
        if get_column_letter(i) == "AZ":
            worksheet.column_dimensions["AZ"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "BA":
            worksheet.column_dimensions["BA"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "BB":
            worksheet.column_dimensions["BB"].width = 12
        if get_column_letter(i) == "BC":
            worksheet.column_dimensions["BC"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
        if get_column_letter(i) == "BD":
            worksheet.column_dimensions["BD"].width = 12
            for row in range(3, len(df_merged_WAGIMS) + 3):
                worksheet.cell(row=row,column=i).number_format = format_6dp
                
                
  aggregate = input("Do you want to calculate sample-level statistics (Y/N):")              
  
  if aggregate.upper() == "Y":
    
    sample_level_name = input("Sample-level file name (include extension):")
    sample_level = os.path.join(out_path,sample_level_name)
                  
    ###### Aggregating functions
    df_igneous = df_merged.loc[df_merged.GroupID == "I"]
    df_grouped = df_igneous.groupby("Sample")
    
    
    variables = ["epsilon_Hf", "TDM_2Stage", "TCR"]
    uncertainties = ["epsilon_Hf_1sig", "TDM_2Stage_5pctUnc", "TCR_5pctUnc"]
    prepends = ["EHFI", "TDM_2Stage_Vervoort2018", "TCR"]
    
    
    while True:
      by_list = input("Type 1 to enter spot numbers manually\nType 2 to read spot list file\n")
      if by_list in ("1","2"):
        break
      else:
        print("Invalid option. Select 1 or 2.")
      
    if by_list == "1":
      composite = create_aggregate_df(df_igneous,"Sample",variables,uncertainties,prepends)
    elif by_list == "2":
      composite = create_aggregate_df(df_igneous,"Sample",variables,uncertainties,prepends,by_list = "exclude.txt")
    
    
                  
                
    with pd.ExcelWriter(sample_level, engine="openpyxl") as writer:
      composite.to_excel(writer, sheet_name="Sample-level stats")
  
      workbook  = writer.book
      worksheet = writer.sheets["Sample-level stats"]
  
      ######### row heights
  
      for cell in worksheet[1]:
          cell.alignment = Alignment(wrap_text=True)
  
      ######### column widths
      for i, col in enumerate(df_merged_WAGIMS.columns, 2):
        
        if get_column_letter(i) in ["B","C","D","E","F","G","H","I"]:
          worksheet.column_dimensions[get_column_letter(i)].width = 10
          for row in range(2, len(composite) + 2):
            worksheet.cell(row=row,column=i).number_format = format_2dp
        elif get_column_letter(i) in ["Z","AA"]:
          worksheet.column_dimensions[get_column_letter(i)].width = 10
          for row in range(2, len(composite) + 2):
            worksheet.cell(row=row,column=i).number_format = format_1dp
        else:
          worksheet.column_dimensions[get_column_letter(i)].width = 10
          for row in range(2, len(composite) + 2):
            worksheet.cell(row=row,column=i).number_format = "0"
            
  else:
      pass

      
              
  

if __name__ == "__main__":
  main()

